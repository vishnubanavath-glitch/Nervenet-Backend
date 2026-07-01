from mcp.server.fastmcp import FastMCP
import openpyxl
from pathlib import Path
import time
import os

mcp = FastMCP("Electricity Meter Server")

DATA_FILE = Path(__file__).parent.parent / "tpcodl_Test.xlsx"
if not DATA_FILE.exists():
    DATA_FILE = Path(__file__).parent / "tpcodl_Test.xlsx"

def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s

class ExcelRepository:
    """
    Responsible only for direct reading and writing of the Excel database file.
    Does not contain any MCP or querying business logic.
    """
    def __init__(self, data_file: Path):
        self.data_file = data_file

    def load_records(self) -> tuple[list[str], list[dict]]:
        """
        Loads workbook from disk and parses it into headers and a list of dictionary records.
        """
        if not self.data_file.exists():
            return [], []
        wb = openpyxl.load_workbook(self.data_file, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []
        
        headers = [str(h) for h in rows[0]]
        records = []
        for row_vals in rows[1:]:
            if any(v is not None for v in row_vals):
                record = dict(zip(headers, row_vals))
                # Normalize primary/indexed fields
                if "uidNo" in record:
                    record["uidNo"] = clean_str(record["uidNo"])
                if "mobileNo" in record:
                    record["mobileNo"] = clean_str(record["mobileNo"])
                records.append(record)
        return headers, records


    def insert_record(self, record: dict):
        """
        Appends a new record to the Excel spreadsheet.
        """
        wb = openpyxl.load_workbook(self.data_file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        new_row = [record.get(h) for h in headers]
        sheet.append(new_row)
        wb.save(self.data_file)

    def update_record(self, uid: str, updates: dict):
        """
        Locates a record by uidNo and updates the specified cells.
        """
        wb = openpyxl.load_workbook(self.data_file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        if "uidNo" not in headers:
            raise ValueError("Database configuration error: uidNo column not found.")
        
        uid_col_idx = headers.index("uidNo") + 1
        row_idx_to_update = None
        
        for row_idx in range(2, sheet.max_row + 1):
            cell_val = clean_str(sheet.cell(row=row_idx, column=uid_col_idx).value)
            if cell_val == uid:
                row_idx_to_update = row_idx
                break
                
        if row_idx_to_update is None:
            raise ValueError(f"Record with uidNo '{uid}' not found in database.")
            
        for key, val in updates.items():
            if key in headers:
                col_idx = headers.index(key) + 1
                sheet.cell(row=row_idx_to_update, column=col_idx, value=val)
                
        wb.save(self.data_file)

    def delete_record(self, uid: str):
        """
        Deletes the row matching the given uidNo from the Excel spreadsheet.
        """
        wb = openpyxl.load_workbook(self.data_file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        if "uidNo" not in headers:
            raise ValueError("Database configuration error: uidNo column not found.")
            
        uid_col_idx = headers.index("uidNo") + 1
        row_idx_to_delete = None
        
        for row_idx in range(2, sheet.max_row + 1):
            cell_val = clean_str(sheet.cell(row=row_idx, column=uid_col_idx).value)
            if cell_val == uid:
                row_idx_to_delete = row_idx
                break
                
        if row_idx_to_delete is None:
            raise ValueError(f"Record with uidNo '{uid}' not found in database.")
            
        sheet.delete_rows(row_idx_to_delete, 1)
        wb.save(self.data_file)


class QueryService:
    """
    Handles in-memory caching, indexing, data projections, filtering,
    sorting, paginations, and aggregations. Contains all system business logic.
    """
    def __init__(self, repository: ExcelRepository):
        self.repository = repository
        self.headers = []
        self.records = []
        # Predefined indexes mapping value.lower() -> list of records containing that value
        self.indexes = {
            "uidNo": {},
            "mobileNo": {},
            "readingStatus": {},
            "subDiv": {}
        }
        self.reload_cache()

    def reload_cache(self):
        """
        Loads data from repository and rebuilds query indexes.
        """
        self.headers, self.records = self.repository.load_records()
        self.rebuild_indexes()

    def rebuild_indexes(self):
        """
        Rebuilds in-memory lookup indexes for fast querying.
        """
        for field in self.indexes:
            self.indexes[field] = {}
            
        for r in self.records:
            for field in self.indexes:
                val = clean_str(r.get(field))
                if val:
                    val_lower = val.lower()
                    if val_lower not in self.indexes[field]:
                        self.indexes[field][val_lower] = []
                    self.indexes[field][val_lower].append(r)

    def query(self, filters: list = None, columns: list = None, sort_by: str = None, descending: bool = False, limit: int = None, offset: int = 0) -> list[dict]:
        """
        Executes query with filtering, sorting, pagination, and projection selection.
        """
        # Convert dictionary filters to list of dicts for backward compatibility
        if isinstance(filters, dict):
            converted = []
            for k, val in filters.items():
                op = "=" if k in ["uidNo", "mobileNo"] else "contains"
                converted.append({"column": k, "operator": op, "value": val})
            filters = converted

        # Validate requested columns
        if columns:
            for col in columns:
                if col not in self.headers:
                    raise ValueError(f"Unknown projection column: '{col}'")

        # Validate structured filters
        if filters:
            if not isinstance(filters, list):
                raise ValueError("Filters must be a list of filter definitions.")
            for f in filters:
                if not isinstance(f, dict) or "column" not in f or "operator" not in f or "value" not in f:
                    raise ValueError("Each filter must be a dictionary with 'column', 'operator', and 'value' keys.")
                col = f.get("column")
                if col not in self.headers:
                    raise ValueError(f"Unknown filter column: '{col}'")
                
                # Check valid operators
                valid_ops = ["=", "!=", ">", "<", ">=", "<=", "contains", "starts_with", "ends_with", "in"]
                op = str(f.get("operator")).strip().lower()
                if op not in valid_ops:
                    raise ValueError(f"Unsupported operator '{op}'. Supported operators: {valid_ops}")

        # Helper method for comparisons
        def _evaluate_op(r_val, operator, f_val) -> bool:
            op = str(operator).strip().lower()
            
            if op == "in":
                if isinstance(f_val, (list, tuple, set)):
                    options = [str(x).strip().lower() for x in f_val]
                else:
                    options = [str(x).strip().lower() for x in str(f_val).split(",")]
                return str(r_val).strip().lower() in options

            try:
                # Numeric comparison
                r_num = float(r_val)
                f_num = float(f_val)
                if op == "=": return r_num == f_num
                elif op == "!=": return r_num != f_num
                elif op == ">": return r_num > f_num
                elif op == "<": return r_num < f_num
                elif op == ">=": return r_num >= f_num
                elif op == "<=": return r_num <= f_num
            except (ValueError, TypeError):
                pass

            # String case-insensitive comparison
            r_str = str(r_val).strip().lower()
            f_str = str(f_val).strip().lower()
            
            if op == "=": return r_str == f_str
            elif op == "!=": return r_str != f_str
            elif op == "contains": return f_str in r_str
            elif op == "starts_with": return r_str.startswith(f_str)
            elif op == "ends_with": return r_str.endswith(f_str)
            
            if op == ">": return r_str > f_str
            elif op == "<": return r_str < f_str
            elif op == ">=": return r_str >= f_str
            elif op == "<=": return r_str <= f_str

            return False

        # Select target records pool
        candidates = self.records
        used_index = False
        indexed_filter = None
        
        # Optimize using index if we have exact match filter on an indexed column
        if filters:
            for f in filters:
                col = f.get("column")
                op = str(f.get("operator")).strip().lower()
                val = f.get("value")
                if col in ["uidNo", "mobileNo", "readingStatus", "subDiv"] and op in ["=", "eq"]:
                    val_lower = clean_str(val).lower()
                    candidates = self.indexes[col].get(val_lower, [])
                    used_index = True
                    indexed_filter = f
                    break

        # Filter candidates
        filtered_records = []
        for r in candidates:
            match = True
            if filters:
                for f in filters:
                    if used_index and f is indexed_filter:
                        continue
                    r_val = r.get(f.get("column"))
                    op = f.get("operator")
                    f_val = f.get("value")
                    
                    if not _evaluate_op(r_val, op, f_val):
                        match = False
                        break
            if match:
                filtered_records.append(r)

        # Sort results
        if sort_by:
            if sort_by not in self.headers:
                raise ValueError(f"Unknown sorting column: '{sort_by}'")
                
            def get_sort_key(record):
                val = record.get(sort_by)
                if val is None:
                    return ""
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return str(val).lower()
                    
            filtered_records.sort(key=get_sort_key, reverse=descending)

        # Paginate results
        start = offset
        end = offset + limit if limit is not None else len(filtered_records)
        paginated_records = filtered_records[start:end]

        # Apply projections (reduce returned columns)
        if columns:
            projected_records = []
            for r in paginated_records:
                projected_records.append({col: r.get(col) for col in columns})
            return projected_records

        return paginated_records

    def prepare_chart_data(self, chart_type: str, title: str, x_axis: str, y_axis: str, aggregation: str = None, filters: list = None, sort_by: str = None, descending: bool = False, limit: int = None) -> dict:
        """
        Processes and aggregates data internally to prepare a clean, chart-ready payload.
        """
        if x_axis not in self.headers:
            raise ValueError(f"Unknown X-axis column: '{x_axis}'")
        if y_axis not in self.headers:
            raise ValueError(f"Unknown Y-axis column: '{y_axis}'")

        records = self.query(filters=filters)

        # Handle group-by aggregation
        if aggregation:
            valid_ops = ["count", "sum", "average", "minimum", "maximum", "median"]
            op_lower = aggregation.lower()
            if op_lower not in valid_ops:
                raise ValueError(f"Unsupported aggregation operation '{aggregation}'. Supported: {valid_ops}")

            groups = {}
            for r in records:
                x_val = clean_str(r.get(x_axis))
                if x_val == "":
                    x_val = "Missing/Empty"
                
                y_val = r.get(y_axis)
                if y_val is not None:
                    if x_val not in groups:
                        groups[x_val] = []
                    groups[x_val].append(y_val)

            chart_data = []
            for x_val, y_vals in groups.items():
                if op_lower == "count":
                    result = len(y_vals)
                else:
                    numeric_y_vals = []
                    for v in y_vals:
                        try:
                            numeric_y_vals.append(float(v))
                        except (ValueError, TypeError):
                            continue

                    if not numeric_y_vals:
                        result = 0.0
                    elif op_lower == "sum":
                        result = sum(numeric_y_vals)
                    elif op_lower == "average":
                        result = sum(numeric_y_vals) / len(numeric_y_vals)
                    elif op_lower == "minimum":
                        result = min(numeric_y_vals)
                    elif op_lower == "maximum":
                        result = max(numeric_y_vals)
                    elif op_lower == "median":
                        numeric_y_vals.sort()
                        n = len(numeric_y_vals)
                        if n % 2 == 1:
                            result = numeric_y_vals[n // 2]
                        else:
                            result = (numeric_y_vals[n // 2 - 1] + numeric_y_vals[n // 2]) / 2.0
                
                chart_data.append({
                    x_axis: x_val,
                    y_axis: result
                })
        else:
            chart_data = []
            for r in records:
                y_val = r.get(y_axis)
                try:
                    y_val = float(y_val) if y_val is not None else 0.0
                except (ValueError, TypeError):
                    pass
                
                chart_data.append({
                    x_axis: clean_str(r.get(x_axis)),
                    y_axis: y_val
                })

        # Sorting
        if sort_by:
            if sort_by not in [x_axis, y_axis]:
                raise ValueError(f"Sort column must be X-axis ('{x_axis}') or Y-axis ('{y_axis}') field.")
            
            def get_sort_key(item):
                val = item.get(sort_by)
                if val is None:
                    return ""
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return str(val).lower()
            chart_data.sort(key=get_sort_key, reverse=descending)
        elif aggregation:
            def get_sort_key(item):
                try:
                    return float(item.get(y_axis))
                except (ValueError, TypeError):
                    return 0.0
            chart_data.sort(key=get_sort_key, reverse=descending)

        if limit is not None:
            chart_data = chart_data[:limit]

        return {
            "type": chart_type.lower().strip(),
            "title": title,
            "x_axis": x_axis,
            "y_axis": y_axis,
            "data": chart_data
        }

    def aggregate(self, operation: str, column: str, filters: list = None):
        """
        Performs in-memory data calculations.
        """
        if column not in self.headers:
            raise ValueError(f"Unknown aggregation column: '{column}'")
            
        valid_ops = ["count", "sum", "average", "minimum", "maximum", "median", "distinct count"]
        op_lower = operation.lower()
        if op_lower not in valid_ops:
            raise ValueError(f"Unsupported aggregation operation: '{operation}'. Supported operations: {valid_ops}")

        matched = self.query(filters=filters)
        if not matched:
            if op_lower in ["count", "distinct count"]:
                return 0
            return None

        # Extract values
        values = [r.get(column) for r in matched if r.get(column) is not None]

        if op_lower == "count":
            return len(values)
        if op_lower == "distinct count":
            return len(set(values))

        # Numeric conversions
        numeric_values = []
        for v in values:
            try:
                numeric_values.append(float(v))
            except (ValueError, TypeError):
                continue

        if not numeric_values:
            return None

        if op_lower == "sum":
            return sum(numeric_values)
        elif op_lower == "average":
            return sum(numeric_values) / len(numeric_values)
        elif op_lower == "minimum":
            return min(numeric_values)
        elif op_lower == "maximum":
            return max(numeric_values)
        elif op_lower == "median":
            numeric_values.sort()
            n = len(numeric_values)
            if n % 2 == 1:
                return numeric_values[n // 2]
            else:
                return (numeric_values[n // 2 - 1] + numeric_values[n // 2]) / 2.0

    def search(self, query_text: str, fields: list = None, limit: int = 5) -> list[dict]:
        """
        Executes free-text substring search across configured searchable fields.
        """
        if not query_text:
            return []
            
        if fields:
            for f in fields:
                if f not in self.headers:
                    raise ValueError(f"Unknown search field: '{f}'")
        else:
            fields = self.headers

        query_lower = query_text.lower()
        results = []
        
        for r in self.records:
            for f in fields:
                val_str = clean_str(r.get(f)).lower()
                if query_lower in val_str:
                    results.append(r)
                    break
            if len(results) >= limit:
                break
                
        return results

    def statistics(self) -> dict:
        """
        Compiles dataset structure statistics.
        """
        total_rows = len(self.records)
        col_count = len(self.headers)
        
        # Duplicate records count
        seen = set()
        duplicates = 0
        for r in self.records:
            rep = tuple(clean_str(r.get(h)) for h in self.headers)
            if rep in seen:
                duplicates += 1
            else:
                seen.add(rep)

        # Missing values per column
        missing_values = {h: 0 for h in self.headers}
        for r in self.records:
            for h in self.headers:
                val = r.get(h)
                if val is None or str(val).strip() == "":
                    missing_values[h] += 1
                    
        # File modification time
        last_modified = None
        try:
            mtime = self.repository.data_file.stat().st_mtime
            last_modified = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mtime))
        except Exception:
            pass

        return {
            "total_rows": total_rows,
            "column_count": col_count,
            "duplicate_rows": duplicates,
            "missing_values": missing_values,
            "last_modified": last_modified
        }


# Initialize components
repository = ExcelRepository(DATA_FILE)
service = QueryService(repository)


# ==========================================
# MCP TOOLS (Validation & Interface Layer)
# ==========================================

@mcp.tool()
def discover_schema():
    """
    Returns high-level structural information about the dataset.
    Use this first to understand available column headers, data types, searchable columns, and aggregatable columns.
    """
    try:
        return service.discover_schema()
    except Exception as e:
        return {"error": f"Failed to discover schema: {str(e)}"}

# Add helper method to QueryService for discover_schema mapping
def discover_schema_method(self):
    col_names = self.headers
    detected_types = {}
    for col in col_names:
        sample_vals = []
        for r in self.records:
            val = r.get(col)
            if val is not None and str(val).strip() != "":
                sample_vals.append(val)
                if len(sample_vals) >= 100:
                    break
        if not sample_vals:
            detected_types[col] = "empty/unknown"
            continue
            
        is_numeric = True
        for v in sample_vals:
            try:
                float(v)
            except (ValueError, TypeError):
                is_numeric = False
                break
        detected_types[col] = "numeric" if is_numeric else "string"

    return {
        "dataset_name": self.repository.data_file.name,
        "total_rows": len(self.records),
        "total_columns": len(col_names),
        "column_names": col_names,
        "detected_data_types": detected_types,
        "searchable_columns": col_names,
        "aggregatable_columns": [col for col, dtype in detected_types.items() if dtype == "numeric"]
    }

# Bind dynamically to QueryService
QueryService.discover_schema = discover_schema_method


@mcp.tool()
def query(filters: list = None, columns: list = None, sort_by: str = None, descending: bool = False, limit: int = None, offset: int = 0):
    """
    Queries customer records applying filter operations, sorting options, pagination, and projection limits.
    - filters: A list of dicts. E.g. [{"column": "step_count", "operator": ">", "value": 125}]
    - columns: Projections subset. Specify only required columns to minimize response size. E.g. ["uidNo", "mobileNo"]
    - sort_by: Sorting column header name. E.g. "step_count"
    - descending: Set to True for descending sort order.
    - limit: Max records returned (default returns all matches).
    - offset: Zero-indexed pagination start offset.
    """
    try:
        res = service.query(filters, columns, sort_by, descending, limit, offset)
        return {"count": len(res), "records": res}
    except Exception as e:
        return {"error": f"Query execution failed: {str(e)}"}


@mcp.tool()
def aggregate(operation: str, column: str, filters: list = None):
    """
    Executes in-memory calculations (count, sum, average, minimum, maximum, median, distinct count).
    - operation: One of: "count", "sum", "average", "minimum", "maximum", "median", "distinct count".
    - column: Column header name to aggregate.
    - filters: Optional filters applying prior to calculation.
    """
    try:
        res = service.aggregate(operation, column, filters)
        return {
            "operation": operation,
            "column": column,
            "result": res
        }
    except Exception as e:
        return {"error": f"Aggregation failed: {str(e)}"}


@mcp.tool()
def search(query_text: str, fields: list = None, limit: int = 5):
    """
    Searches keyword substrings across multiple database columns.
    - query_text: The search text query. E.g. "KWH" or subdivision name.
    - fields: Optional columns subset to match against.
    - limit: Maximum matches to return (default 5).
    """
    try:
        res = service.search(query_text, fields, limit)
        return {"count": len(res), "records": res}
    except Exception as e:
        return {"error": f"Search failed: {str(e)}"}


@mcp.tool()
def statistics():
    """
    Compiles dataset metadata overview (total rows, missing value distribution, duplicates count).
    """
    try:
        return service.statistics()
    except Exception as e:
        return {"error": f"Failed compiling statistics: {str(e)}"}


@mcp.tool()
def sample_rows(limit: int = 5):
    """
    Returns a small sample row slice to inspect cell contents (default 5 rows).
    """
    try:
        size = min(limit, 50)
        res = service.query(limit=size)
        return {"sample_size": len(res), "records": res}
    except Exception as e:
        return {"error": f"Failed fetching sample: {str(e)}"}


@mcp.tool()
def create(record: dict):
    """
    Inserts a new customer record into the spreadsheet database.
    - record: Dictionary containing headers and values. uidNo is required and must be unique.
    """
    try:
        # Check required primary key
        uid = clean_str(record.get("uidNo"))
        if not uid:
            return {"error": "Missing required field: 'uidNo'"}
            
        # Check database columns validation
        for col in record.keys():
            if col not in service.headers:
                return {"error": f"Validation failed: Column '{col}' does not exist in database schema."}
                
        # Check duplication
        if uid.lower() in service.indexes["uidNo"]:
            return {"error": f"Duplicate primary key error: A record with uidNo '{uid}' already exists."}
            
        repository.insert_record(record)
        service.reload_cache()
        return {"status": "Success", "message": f"Customer record for uidNo '{uid}' inserted successfully."}
    except Exception as e:
        return {"error": f"Failed inserting record: {str(e)}"}


@mcp.tool()
def update(uid: str, updates: dict):
    """
    Updates record properties inside the database.
    - uid: The target unique uidNo to modify.
    - updates: Dict containing target headers and updated values.
    """
    try:
        uid_clean = clean_str(uid)
        if not uid_clean:
            return {"error": "Missing parameter 'uid'"}
            
        if uid_clean.lower() not in service.indexes["uidNo"]:
            return {"error": f"Record with uidNo '{uid_clean}' not found."}
            
        for col in updates.keys():
            if col not in service.headers:
                return {"error": f"Validation failed: Column '{col}' does not exist in database schema."}
                
        if "uidNo" in updates and clean_str(updates["uidNo"]) != uid_clean:
            return {"error": "Modifying uidNo primary key value is prohibited during updates."}
            
        repository.update_record(uid_clean, updates)
        service.reload_cache()
        
        # Return updated record
        updated_rec = service.query(filters={"uidNo": uid_clean}, limit=1)
        return {"status": "Success", "record": updated_rec[0] if updated_rec else None}
    except Exception as e:
        return {"error": f"Update failed: {str(e)}"}


@mcp.tool()
def delete(uid: str):
    """
    Removes a record from the database.
    - uid: Unique uidNo to delete.
    """
    try:
        uid_clean = clean_str(uid)
        if not uid_clean:
            return {"error": "Missing parameter 'uid'"}
            
        if uid_clean.lower() not in service.indexes["uidNo"]:
            return {"error": f"Record with uidNo '{uid_clean}' not found."}
            
        repository.delete_record(uid_clean)
        service.reload_cache()
        return {"status": "Success", "message": f"Record '{uid_clean}' successfully deleted."}
    except Exception as e:
        return {"error": f"Deletion failed: {str(e)}"}


@mcp.tool()
def reload_cache():
    """
    Forces database reload into cache memory, rebuilding lookup indexes.
    """
    try:
        service.reload_cache()
        return {"status": "Success", "message": "Database cache and query indexes successfully reloaded."}
    except Exception as e:
        return {"error": f"Cache reload failed: {str(e)}"}


@mcp.tool()
def prepare_chart_data(chart_type: str, title: str, x_axis: str, y_axis: str, aggregation: str = None, filters: list = None, sort_by: str = None, descending: bool = False, limit: int = None):
    """
    Prepares raw/aggregated customer records for visual representation.
    Returns a standardized JSON dataset structure without rendering the chart itself.
    - chart_type: E.g. "bar", "line", "pie", "scatter", "histogram", "area"
    - title: Description title of the chart E.g. "Top 5 Customers by Step Count"
    - x_axis: Column name to map on the X-Axis (group-by key when aggregating).
    - y_axis: Column name to map on the Y-Axis (value key).
    - aggregation: Optional calculations. One of: "count", "sum", "average", "minimum", "maximum", "median"
    - filters: Optional operator-based filters applying before calculation.
    - sort_by: Field to sort results by. Must be x_axis or y_axis.
    - descending: Set to True for descending sort order.
    - limit: Max records/groups to return (Top-N).
    """
    try:
        return service.prepare_chart_data(chart_type, title, x_axis, y_axis, aggregation, filters, sort_by, descending, limit)
    except Exception as e:
        return {"error": f"Failed preparing chart data: {str(e)}"}


if __name__ == "__main__":
    mcp.run()